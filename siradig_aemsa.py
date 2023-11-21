from bs4 import BeautifulSoup
import openpyxl
import os
import datetime

nombre_libro = '/home/sebastian/Documentos/DriveEstudio/Trabajo/AEMSA/2023/Sueldos/siradig_aemsa_2023_07.xlsx'
etiquetaEmpleado = ['CUIT', 'TipoDoc', 'Apellido', 'Nombre']
etiquetaDireccion = ['Provincia', 'Cod Postal', 'Localidad', 'Calle', 'Número', 'Piso', 'Departamento']
etiquetafamiliar = \
    ['TipoDoc', 'NroDoc', 'Apellido', 'Nombre', 'FechaNacimiento',
     'MesDesde', 'MesHasta', 'Parentesco', 'VigenteProx', 'Porcentaje']
etiquetaDeducciones = ['TipoDeduccion', 'TipoDocumento', 'NroDoc', 'Denominacion', 'DescripcionBasica', 'Monto']
etiquetaRetenciones = ['TipoRetPer', 'TipoDocumento', 'NroDoc', 'Denominacion', 'DescripcionBasica', 'Monto']
etiquetaRemOtroEmpl = \
    ['Mes', 'obraSoc', 'segSoc', 'sind', 'ganBrut', 'retGan', 'retribNoHab', 'ajuste',
     'exeNoAlc', 'sac', 'horasExtGr', 'horasExtEx', 'matDid', 'gastosMovViat', 'bonosProd',
     'fallosCaja', 'conSimNat', 'remunExentaLey27549', 'suplemParticLey19101', 'teletrabajoExento']

dirXML = '/home/sebastian/Documentos/DriveEstudio/Trabajo/AEMSA/2023/Sueldos/siradig/xml'

for file in os.listdir(dirXML):

    with open(f'{dirXML}/{file}', 'r') as f:
        data = f.read()

    Bs_data = BeautifulSoup(data, 'xml')
    nombre_hoja = Bs_data.empleado.apellido.text + ', ' + Bs_data.empleado.nombre.text

    if not os.path.isfile(nombre_libro):
        wb = openpyxl.Workbook()
        ws = wb['Sheet']
        ws.title = nombre_hoja

    else:
        wb = openpyxl.load_workbook(nombre_libro)
        if not nombre_hoja in wb.sheetnames:
            ws = wb.create_sheet(nombre_hoja)
        else:
            ws = wb[nombre_hoja]

    ws.delete_cols(1, 20)

    if not Bs_data.find('periodo') is None:
        for periodo in Bs_data.find('periodo'):
            ws['A1'] = "Período:"
            ws['B1'] = periodo
    else:
        ws['A1'] = "No hay datos del período"


    if not Bs_data.find('nroPresentacion') is None:
        for presPeriodo in Bs_data.find('nroPresentacion'):
            ws['A2'] = "Nro Presentacion:"
            ws['B2'] = presPeriodo

    else:
        ws['A2'] = "No hay datos de la presentacion"

    if not Bs_data.find('fechaPresentacion') is None:
        for fechaPresentacion in Bs_data.find('fechaPresentacion'):
            ws['A3'] = "Fecha Presentacion:"
            ws['B3'] = datetime.datetime.strptime(fechaPresentacion, '%Y-%m-%d')
            ws['B3'].number_format = 'DD/MM/YYYY'
    else:
        ws['A3'] = "No hay fecha de presentación"

    if not Bs_data.find('empleado') is None:
        i, j = ws.max_row + 2, ws.max_row + 2
        for celda in etiquetaEmpleado:
            ws.cell(column=1, row=i).value = celda
            i += 1

        for data in Bs_data.find('empleado'):
            if not data.name == 'direccion':
                if not data.text.strip() == '':
                    ws.cell(column=2, row=j).value = data.text
                    j += 1
            else:
                break

        i, j = ws.max_row + 2, 1
        for data in etiquetaDireccion:
            ws.cell(column=j, row=i).value = data
            j += 1

        i, j = ws.max_row + 1, 1

        for data in Bs_data.select('direccion')[0].text.split('\n'):
            if not data == "":
                ws.cell(column=j, row=i).value = data
                j += 1


    if not Bs_data.find('cargasFamilia') is None:
        i, j = ws.max_row + 2, 1

        for celda in etiquetafamiliar:
            ws.cell(column=j, row=i).value = celda
            j += 1

        i += 1
        for cargasFamilia in Bs_data.select('cargasFamilia'):
            for familiar in cargasFamilia.find_all('cargaFamilia'):
                ws.cell(column=1, row=i).value = familiar.tipoDoc.text
                ws.cell(column=2, row=i).value = familiar.nroDoc.text
                ws.cell(column=3, row=i).value = familiar.apellido.text
                ws.cell(column=4, row=i).value = familiar.nombre.text
                ws.cell(column=5, row=i).value = datetime.datetime.strptime(familiar.fechaNac.text, '%Y-%m-%d')
                ws.cell(column=5, row=i).number_format = 'DD/MM/YYYY'
                ws.cell(column=6, row=i).value = familiar.mesDesde.text
                ws.cell(column=7, row=i).value = familiar.mesHasta.text
                ws.cell(column=8, row=i).value = familiar.parentesco.text
                ws.cell(column=9, row=i).value = familiar.vigenteProximosPeriodos.text
                ws.cell(column=10, row=i).value = familiar.porcentajeDeduccion.text
                i += 1

    else:
        i = ws.max_row + 2
        ws.cell(column=1, row=i).value = "No hay cargas de familia"

    if not Bs_data.find('ganLiqOtrosEmpEnt') is None:
        i, j = ws.max_row + 2, 1
        for ganLiqOtrEmpl in Bs_data.select('ganLiqOtrosEmpEnt'):
            for data in ganLiqOtrEmpl.find_all('empEnt'):
                ws.cell(column=j, row=i).value = data.cuit.text
                i += 1
                ws.cell(column=j, row=i).value = data.denominacion.text
                i += 2
                for celda in etiquetaRemOtroEmpl:
                    ws.cell(column=j, row=i).value = celda
                    j += 1

                i += 1
                for d in data.ingresosAportes.find_all('ingAp'):
                    ws.cell(column=1, row=i).value = d['mes']
                    ws.cell(column=2, row=i).value = float(d.obraSoc.text)
                    ws.cell(column=3, row=i).value = float(d.segSoc.text)
                    ws.cell(column=4, row=i).value = float(d.sind.text)
                    ws.cell(column=5, row=i).value = float(d.ganBrut.text)
                    ws.cell(column=6, row=i).value = float(d.retGan.text)
                    ws.cell(column=7, row=i).value = float(d.retribNoHab.text)
                    ws.cell(column=8, row=i).value = float(d.ajuste.text)
                    ws.cell(column=9, row=i).value = float(d.exeNoAlc.text)
                    ws.cell(column=10, row=i).value = float(d.sac.text)
                    ws.cell(column=11, row=i).value = float(d.horasExtGr.text)
                    ws.cell(column=12, row=i).value = float(d.horasExtEx.text)
                    ws.cell(column=13, row=i).value = float(d.matDid.text)
                    ws.cell(column=14, row=i).value = float(d.gastosMovViat.text)
                    ws.cell(column=15, row=i).value = float(d.bonosProd.text)
                    ws.cell(column=16, row=i).value = float(d.fallosCaja.text)
                    ws.cell(column=17, row=i).value = float(d.conSimNat.text)
                    ws.cell(column=18, row=i).value = float(d.remunExentaLey27549.text)
                    ws.cell(column=19, row=i).value = float(d.suplemParticLey19101.text)
                    ws.cell(column=20, row=i).value = float(d.teletrabajoExento.text)
                    i += 1

                j = 1

    else:
        i = ws.max_row + 2
        ws.cell(column=1, row=i).value = "***** No informa Ganancias de otros empleadores *****"

    if not Bs_data.find('deducciones') is None:
        i, j = ws.max_row + 2, 1

        for celda in etiquetaDeducciones:
            ws.cell(column=j, row=i).value = celda
            j += 1

        i += 1
        for deducciones in Bs_data.select('deducciones'):
                for deduccion in deducciones.find_all('deduccion'):
                    ws.cell(column=1, row=i).value = deduccion['tipo']
                    ws.cell(column=2, row=i).value = deduccion.tipoDoc.text
                    ws.cell(column=3, row=i).value = deduccion.nroDoc.text
                    ws.cell(column=4, row=i).value = deduccion.denominacion.text
                    ws.cell(column=5, row=i).value = deduccion.descBasica.text
                    ws.cell(column=6, row=i).value = float(deduccion.montoTotal.text)
                    i += 1

    else:
        i = ws.max_row + 2
        ws.cell(column=1, row=i).value = "No hay deducciones"


    if not Bs_data.find('retPerPagos') is None:
        i, j = ws.max_row + 2, 1

        for celda in etiquetaRetenciones:
            ws.cell(column=j, row=i).value = celda
            j += 1

        i += 1
        for retencionesPercepciones in Bs_data.select('retPerPagos'):
            for retencion in retencionesPercepciones.find_all('retPerPago'):
                ws.cell(column=1, row=i).value = retencion['tipo']
                ws.cell(column=2, row=i).value = retencion.tipoDoc.text
                ws.cell(column=3, row=i).value = retencion.nroDoc.text
                ws.cell(column=4, row=i).value = retencion.denominacion.text
                ws.cell(column=5, row=i).value = retencion.descBasica.text
                ws.cell(column=6, row=i).value = float(retencion.montoTotal.text)
                i += 1

    else:
        i = ws.max_row + 2
        ws.cell(column=1, row=i).value = "No hay retenciones / percepciones"



    # print("Ajustes")
    if not Bs_data.find('ajustes') is None:
        for ajustes in Bs_data.find('ajustes'):
            pass
            # print(ajustes)
    else:
        pass
        # print("No hay ajustes")

    # print("Datos adicionales")
    if not Bs_data.find('datosAdicionales') is None:
        for datAdicinales in Bs_data.find('datosAdicionales'):
            pass
            # print(datAdicinales)
    else:
        pass
        # print("No hay datos adiciones")

    wb.save(nombre_libro)
    wb.close()

exit(0)