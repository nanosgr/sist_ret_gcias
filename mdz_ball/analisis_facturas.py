# Descripción: Módulo para analizar las facturas de la empresa MDZ Ballon
import openpyxl 
import os
import datetime

def analisis_facturacion(fecha_desde: datetime, fecha_hasta: datetime, nombre_archivo: str):
    # print(f'Vamos a analizar desde el {fecha_desde} hasta el {fecha_hasta}')

    wb = openpyxl.load_workbook(nombre_archivo)
    wsFacturacion = wb['Facturacion']
    wsPasajeros = wb['Pasajeros']

    for i in range(2, wsFacturacion.max_row + 1):
        if wsFacturacion.cell(column=1, row=i).value is None:
            break
        cbte = dict(
            comprobante=wsFacturacion.cell(column=1, row=i).value,
            cantidad_pasajeros=int(wsFacturacion.cell(column=7, row=i).value),
            precio_total=float(wsFacturacion.cell(column=8, row=i).value),
            efectivo=float(wsFacturacion.cell(column=9, row=i).value),
            transferencia=float(wsFacturacion.cell(column=10, row=i).value),
            frances=float(wsFacturacion.cell(column=11, row=i).value),
            mercado_pago=float(wsFacturacion.cell(column=12, row=i).value),
            px_unitario=float(wsFacturacion.cell(column=6, row=i).value))
        analisis_pasajeros(cbte, wsPasajeros)

    wb.save(nombre_archivo)
    wb.close()
    return

def analisis_pasajeros(comprobante: dict, wsPasajeros: openpyxl.worksheet):
    medios_pago = ['efectivo', 'transferencia', 'frances', 'mercado_pago']

    for i in range(2, wsPasajeros.max_row + 1):
        comprobante_pasajero = wsPasajeros.cell(column=11, row=i).value
        pago_apropiado = comprobante['px_unitario']
        if not comprobante_pasajero.find(comprobante['comprobante']) and comprobante['cantidad_pasajeros'] > 0:
            # Con los datos del diccionario comprobante, buscamos para cada uno de los pasajeros 
            # la apropiación de los pagos, y actualizamos la hoja wsPasajeros
            for medio in medios_pago:
                if comprobante[medio] < 0:
                    if (round(comprobante[medio], 0) + comprobante['px_unitario']) <= 0:
                        wsPasajeros.cell(column=medios_pago.index(medio) + 12, row=i).value = comprobante['px_unitario'] if comprobante['px_unitario'] == pago_apropiado else pago_apropiado
                        comprobante[medio] += comprobante['px_unitario'] if comprobante['px_unitario'] == pago_apropiado else pago_apropiado
                        comprobante['cantidad_pasajeros'] -= 1
                        pago_apropiado -= comprobante['px_unitario'] if comprobante['px_unitario'] == pago_apropiado else pago_apropiado
                        break 
                    else:
                        wsPasajeros.cell(column=medios_pago.index(medio) + 12, row=i).value = abs(round(comprobante[medio], 0))
                        pago_apropiado -= abs(round(comprobante[medio], 0))
                        comprobante[medio] = 0
                        if pago_apropiado <= 0:
                            comprobante['cantidad_pasajeros'] -= 1
                            continue

                if pago_apropiado <= 0:
                    break
    
    return