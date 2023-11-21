import openpyxl
import os

nombre_libro = './SIRADIG_AEMSA/siradig_2022.xlsx'

if not os.path.isfile(nombre_libro):
    wb = openpyxl.Workbook()
    ws = wb['Sheet']
    ws.title = nombre_hoja

else:
    wb = openpyxl.load_workbook(nombre_libro)
    if not nombre_hoja in wb.sheetnames:
        ws = wb.create_sheet(nombre_hoja)
    else:
        ws = wb[nombre_hoja]

ws.delete_cols(1, 20)